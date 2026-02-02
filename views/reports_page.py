"""
Reports Page Module
Handles export of call log data to Excel with optional email functionality
"""
import streamlit as st
import io
import smtplib
import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Font
from datetime import datetime
from storage import DateRange
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from time import sleep
from utils import df_from_records

# Main Render Function
def render_reports_page(repo):
    """
    Render the Reports page with data export functionality.
    """
    st.subheader("📊 Export & Analytical Call Log Reports")
    try:
        # 1. Check if any data exists at all
        all_records = repo.calllog_list()
        
        if not all_records:
            st.warning("No data found in the database to export.")
            return

        # 2. Date range filter and Clear Button
        col1, col2, col3 = st.columns([1.5, 1.5, 7])
        with col1:
            start_date = st.date_input("Start Date", value=None, key="start_date")
        with col2:
            end_date = st.date_input("End Date", value=None, key="end_date")
        
        # 3. Apply the filter (None values result in all data being fetched)
        if start_date and end_date:
            dr = DateRange(
                start=datetime.combine(start_date, datetime.min.time()),
                end=datetime.combine(end_date, datetime.max.time()),
            )
        else:
            dr = None
        
        # 4. Get the DataFrame
        df = df_from_records(repo.calllog_list(dr))
        
        # 5. Display table and export options if data is present
        if not df.empty:
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.info(f"Total records: {len(df)}")

            buffer = _create_formatted_excel(df, sheet_name="CallLogReport")
            
            exp_col1, exp_col2 = st.columns([1.5, 8.5])
            with exp_col1:
                st.download_button(
                    label="📥 Download Excel",
                    data=buffer,
                    file_name=f"CallLog_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            with exp_col2:
                _render_email_section(buffer)
        else:
            st.warning("No records found for the selected criteria.")
            
    except Exception as e:
        st.error(f"Error loading reports: {e}")


# Helper Function to render email section
def _render_email_section(buffer):
    """Render email report functionality."""
    repo = st.session_state.active_repo
    
    # Get email config from database
    email_config = repo.email_config_get()
    
    if not email_config:
        st.warning("⚠️ Email not configured. Please configure email settings in the Email page first.")
        return
    
    with st.expander("📧 Email Report"):
        email_to = st.text_input("Recipient Email", placeholder="recipient@example.com")
        email_subject = st.text_input("Subject", value=f"Call Log Report - {datetime.now().strftime('%Y-%m-%d')}")
        email_body = st.text_area("Message", value="Please find the attached call log report.", height=100)
        
        if st.button("Send Email", type="primary"):
            if email_to:
                try:
                    # Use config from database
                    smtp_server = email_config.smtp_server
                    smtp_port = email_config.smtp_port
                    smtp_user = email_config.smtp_user
                    smtp_password = email_config.smtp_password
                    
                    if not smtp_user or not smtp_password:
                        st.error("⚠️ Email configuration incomplete. Please update settings.")
                    else:
                        # Create message
                        msg = MIMEMultipart()
                        msg['From'] = smtp_user
                        msg['To'] = email_to
                        msg['Subject'] = email_subject
                        
                        # Add body
                        msg.attach(MIMEText(email_body, 'plain'))
                        
                        # Attach Excel file
                        buffer.seek(0)
                        filename = f"CallLog_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        part.set_payload(buffer.read())
                        encoders.encode_base64(part)
                        part.add_header('Content-Disposition', f'attachment; filename={filename}')
                        msg.attach(part)
                        
                        # Send email
                        with st.spinner("Sending email...", show_time=True):
                          with smtplib.SMTP(smtp_server, smtp_port) as server:
                              server.starttls()
                              server.login(smtp_user, smtp_password)
                              server.send_message(msg)
                          sleep(3)  # Simulate delay for sending
                        st.success(f"✅ Email was sent successfully to {email_to} !!", icon='📧')
                except Exception as e:
                    st.error(f"❌ Error sending email: {e}")
            else:
                st.warning("Please enter a recipient email address.")


# Helper Function to create formatted Excel
def _create_formatted_excel(df: pd.DataFrame, sheet_name: str = 'CallLogEntries') -> io.BytesIO:
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        
        # 1. Define Styles
        # Navy blue header with white bold text
        header_fill = PatternFill(start_color="1F305E", end_color="1F305E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Light grey fill for alternating rows
        alternate_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        center_style = Alignment(horizontal='center', vertical='center')

        # 2. Apply Styles to Header (Row 1)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_style

        # 3. Apply Styles to Data Rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.alignment = center_style
                # Apply alternate color to even rows
                if row_idx % 2 == 0:
                    cell.fill = alternate_fill

        # 4. Auto-Adjust Column Widths
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = max_length + 5
            
    buffer.seek(0)
    return buffer